#!/usr/bin/env python3
"""Export JSON data to Excel (.xlsx) or CSV with formatted tables."""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, numbers
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def load_json(input_path: str | None) -> list[dict]:
    """Load JSON data from file or stdin."""
    if input_path:
        with open(input_path) as f:
            data = json.load(f)
    else:
        data = json.load(sys.stdin)

    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for val in data.values():
            if isinstance(val, list) and val and isinstance(val[0], dict):
                return val
        return [data]
    return []


def extract_sheets(data) -> dict[str, list[dict]]:
    """Extract multiple sheets from nested JSON structure."""
    if isinstance(data, list):
        return {"Sheet1": data}
    if isinstance(data, dict):
        sheets = {}
        for key, val in data.items():
            if isinstance(val, list) and val and isinstance(val[0], dict):
                sheets[key] = val
        if sheets:
            return sheets
        return {"Sheet1": [data]}
    return {"Sheet1": []}


def write_xlsx(sheets: dict[str, list[dict]], output_path: str, sheet_name: str | None) -> None:
    """Write data to Excel with formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active

    if len(sheets) == 1 and sheet_name:
        sheets = {sheet_name: list(sheets.values())[0]}
    elif sheet_name and "Sheet1" in sheets:
        sheets[sheet_name] = sheets.pop("Sheet1")

    first = True
    for name, rows in sheets.items():
        if not rows:
            continue
        if first:
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(title=name)

        headers = list(rows[0].keys())
        bold = Font(bold=True)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold

        for row_idx, row in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                val = row.get(header)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if isinstance(val, float):
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row in rows:
                val_len = len(str(row.get(header, "")))
                if val_len > max_len:
                    max_len = val_len
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

    wb.save(output_path)


def write_csv(rows: list[dict], output_path: str) -> None:
    """Write data to CSV."""
    if not rows:
        print("No data to export.", file=sys.stderr)
        return
    headers = list(rows[0].keys())
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export JSON data to Excel or CSV")
    parser.add_argument("--input", help="JSON input file (reads stdin if omitted)")
    parser.add_argument("--output", required=True, help="Output file (.xlsx or .csv)")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: Sheet1)")
    args = parser.parse_args()

    raw_input = args.input
    if raw_input:
        with open(raw_input) as f:
            raw_data = json.load(f)
    else:
        raw_data = json.load(sys.stdin)

    rows = load_json(args.input) if raw_input else []
    if not rows and isinstance(raw_data, list):
        rows = raw_data
    elif not rows and isinstance(raw_data, dict):
        for val in raw_data.values():
            if isinstance(val, list):
                rows = val
                break
        if not rows:
            rows = [raw_data]

    is_xlsx = args.output.lower().endswith(".xlsx")

    if is_xlsx and HAS_OPENPYXL:
        sheets = extract_sheets(raw_data)
        write_xlsx(sheets, args.output, args.sheet)
        print("Exported to: {}".format(os.path.abspath(args.output)))
    elif is_xlsx and not HAS_OPENPYXL:
        fallback = args.output.rsplit(".", 1)[0] + ".csv"
        print("openpyxl not installed, falling back to CSV.", file=sys.stderr)
        write_csv(rows, fallback)
        print("Exported to: {}".format(os.path.abspath(fallback)))
    else:
        write_csv(rows, args.output)
        print("Exported to: {}".format(os.path.abspath(args.output)))


if __name__ == "__main__":
    main()
